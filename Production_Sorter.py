import pandas as pd
import openpyxl as xl
import os
import sys

downloads_dir = os.path.join(os.path.expanduser("~"), "Downloads")

print("NOTE: Leave your file in your 'Downloads' folder")

def load_csv_file():
    while True:
        file_name = input("Please input your file name (without extension): ")
        file_path = os.path.join(downloads_dir, file_name + ".csv")
        
        try:
            production_df = pd.read_csv(file_path)
            print("File loaded successfully!")
            return production_df, file_name
        except FileNotFoundError:
            print("That file doesn't exist in the Downloads folder. Please try again.")
        except Exception as e:
            print(f"An error occurred: {e}. Please try again.")

exit = False

while not exit:
    cont = input("Would you like to process another file? Y/N: ").upper()
    if cont == "Y":
        production_df, file_name = load_csv_file()
    else:
        sys.exit()

if file_name.find("Production") != -1:
    production_df['TXN_DATE'] = pd.to_datetime(production_df['TXN_DATE'])

    headers = ["PART_NBR", "BIN_ID", "TXN_QTY", "USER NAME", "TXN_DATE", "SUB CODE"]

    filtered_df = production_df[production_df["APPLICATION"] == "PICKING"].copy()

    filtered_df["SERIAL"] = filtered_df.apply(lambda row: f"{row['PART_NBR']}{row['BIN_ID']}{row['USER NAME']}{row['TXN_DATE'].strftime('%Y%m%d%H%M')}{row['SUB CODE']}", axis=1)

    grouped_df = filtered_df.groupby("SERIAL").agg({
        "PART_NBR": "first",
        "BIN_ID": "first",
        "TXN_QTY": "sum",
        "USER NAME": "first",
        "TXN_DATE": "first",
        "SUB CODE": "first"
    }).reset_index()

    grouped_df.drop('SERIAL', axis=1, inplace=True)

    wb = xl.Workbook()
    ws = wb.active

    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col).value = header

    column_mapping = {header: chr(65 + idx) for idx, header in enumerate(headers)}

    date_column_letter = chr(65 + headers.index("TXN_DATE"))
    for row_index, row in grouped_df.iterrows():
        for key, column in column_mapping.items():
            cell_address = f"{column}{row_index + 2}" 
            cell = ws[cell_address]
            cell.value = row[key]
            if key == "TXN_DATE" and isinstance(row[key], pd.Timestamp):
                cell.number_format = 'M/D/YYYY HH:MM' 

    print("Saving File...")
    wb.save(os.path.join(downloads_dir, file_name + ".xlsx"))
    
if "Bin" in file_name:
    production_df['COUNT_DATE'] = pd.to_datetime(production_df['COUNT_DATE'])

    production_df['COUNT_DAY'] = production_df['COUNT_DATE'].dt.date

    headers = ["FACILITY_ID", "BIN_SOURCE", "BUILDING", "BIN_ID", "PART_NBR", "PART_DESC", "SYSTEM_QTY", "COUNT_QTY", "DELTA", "COUNT_DATE", "COUNTED_BY"]

    production_df['SERIAL'] = production_df.apply(lambda row: f"{row['FACILITY_ID']}{row['BIN_SOURCE']}{row['BUILDING']}{row['BIN_ID']}{row['PART_NBR']}{row['SYSTEM_QTY']}{row['COUNT_DATE'].strftime('%Y%m%d%H%M')}{row['COUNTED_BY']}", axis=1)

    grouped_df = production_df.sort_values('COUNT_DATE').groupby(['SERIAL', 'COUNT_DAY']).agg({
        "FACILITY_ID": "last", 
        "BIN_SOURCE": "last", 
        "BUILDING": "last", 
        "BIN_ID": "last", 
        "PART_NBR": "last", 
        "PART_DESC": "last", 
        "SYSTEM_QTY": "last", 
        "COUNT_QTY": "last", 
        "DELTA": "last", 
        "COUNT_DATE": "last", 
        "COUNTED_BY": "last"
    }).reset_index()

    grouped_df.drop(['SERIAL', 'COUNT_DAY'], axis=1, inplace=True)

    wb = xl.Workbook()
    ws = wb.active

    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col).value = header

    column_mapping = {header: chr(65 + idx) for idx, header in enumerate(headers)}

    date_column_letter = chr(65 + headers.index("COUNT_DATE"))
    for row_index, row in grouped_df.iterrows():
        for key, column in column_mapping.items():
            cell_address = f"{column}{row_index + 2}"
            cell = ws[cell_address]
            cell.value = row[key]
            if key == "COUNT_DATE" and isinstance(row[key], pd.Timestamp):
                cell.number_format = 'M/D/YYYY HH:MM'

    print("Saving File...")
    wb.save(os.path.join(downloads_dir, file_name + ".xlsx"))